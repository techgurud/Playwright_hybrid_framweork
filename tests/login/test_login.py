import pytest
import openpyxl
from playwright.sync_api import sync_playwright
from page_objects.login_page import LoginPage
from page_objects.home_page import HomePage

# Path to the test data file
TEST_DATA_FILE = 'testdata.xlsx'

def read_test_data(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    test_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        test_data.append({'username': row[0], 'password': row[1], 'expected_result': row[2]})
    return test_data, workbook, sheet

def write_test_result(sheet, row, result):
    sheet.cell(row=row, column=4, value=result)

@pytest.mark.parametrize("data", read_test_data(TEST_DATA_FILE)[0])
def test_login(data):
    test_data, workbook, sheet = read_test_data(TEST_DATA_FILE)
    row = test_data.index(data) + 2  # Adjust for header row

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context()
        page = context.new_page()

        login_page = LoginPage(page)
        home_page = HomePage(page)

        try:
            # Navigate to the login page
            login_page.navigate()

            # Perform login
            login_page.login(data['username'], data['password'])

            # Validate login result
            if data['expected_result'] == "Pass":
                assert home_page.is_dashboard_displayed()
                write_test_result(sheet, row, "Pass")
            else:
                assert login_page.is_invalid_credentials_displayed()
                write_test_result(sheet, row, "Fail")

        except Exception as e:
            write_test_result(sheet, row, f"Error: {str(e)}")
        finally:
            browser.close()

    # Save the updated Excel file
    workbook.save(TEST_DATA_FILE)
    workbook.close()